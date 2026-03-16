"""
parser_corp2.py — парсер для корпуса 2.

Основной файл: горизонтальная таблица
  Строка 1: названия групп (каждые 3 столбца)
  Строка 2: шапка (Предмет, Преподаватель, Кабинет)
  Строки 3+: дни недели с парами 0-6

Файлы замен: вертикальная таблица (как корпус 3/4)
  Строка 1: заголовок с датой и днём недели
  Далее: группы-заголовки + пары
"""

import io
import re
import logging
from datetime import date, timedelta

import openpyxl

logger = logging.getLogger(__name__)

DAYS_RU_UPPER = {
    'ПОНЕДЕЛЬНИК': 0, 'ВТОРНИК': 1, 'СРЕДА': 2,
    'ЧЕТВЕРГ': 3, 'ПЯТНИЦА': 4, 'СУББОТА': 5, 'ВОСКРЕСЕНЬЕ': 6,
}

DAYS_RU_CAP = {v: k.capitalize() for k, v in DAYS_RU_UPPER.items()}


def _get_week_number(d: date) -> int:
    """
    Определяет номер учебной недели (1 или 2).
    Отсчёт от 1 сентября — нечётные ISO-недели = 1, чётные = 2.
    """
    # Начало учебного года — ближайший понедельник к 1 сентября
    year = d.year if d.month >= 9 else d.year - 1
    sep1 = date(year, 9, 1)
    # Первый понедельник сентября или сам 1 сентября если понедельник
    days_to_mon = (7 - sep1.weekday()) % 7
    first_mon = sep1 if sep1.weekday() == 0 else sep1 + timedelta(days=days_to_mon)
    weeks_passed = (d - first_mon).days // 7
    return 1 if weeks_passed % 2 == 0 else 2


def _parse_ned_cell(cell_val: str, week: int) -> tuple[str, str, str]:
    """
    Разбирает ячейку с НЕД-паттерном.
    '1 НЕД Математика/\n2 НЕД Литература' -> для нужной недели возвращает предмет.
    Возвращает (subject, raw_teacher, raw_room) где teacher и room могут быть пустыми.
    """
    if not cell_val or 'НЕД' not in str(cell_val).upper():
        return str(cell_val).strip() if cell_val else '', '', ''

    text = str(cell_val).strip()
    # Разбиваем по / или \n
    parts = re.split(r'/|\n', text)

    ned1_subj = ''
    ned2_subj = ''

    for part in parts:
        part = part.strip()
        # Ищем паттерн "1 НЕД ..." или "2 НЕД ..."
        m1 = re.match(r'^1\s*[Нн][Еe][Дд][.\s]*(.*)$', part)
        m2 = re.match(r'^2\s*[Нн][Еe][Дд][.\s]*(.*)$', part)
        if m1:
            ned1_subj = m1.group(1).strip().strip('-').strip()
        elif m2:
            ned2_subj = m2.group(1).strip().strip('-').strip()

    subj = ned1_subj if week == 1 else ned2_subj
    return subj, '', ''


def _fmt_val(val) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    # float → int
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (ValueError, AttributeError):
        pass
    return s


def _split_multi_pair(num_str: str) -> list[str]:
    """
    Разбивает "1,2" или "1-2" на ['1', '2'].
    Для обычных номеров возвращает [num_str].
    """
    s = num_str.strip()
    m = re.match(r'^(\d+)[,\-](\d+)$', s)
    if m:
        return [m.group(1), m.group(2)]
    return [s]


# ─── Парсер основного файла ───────────────────────────────────────────────────

def parse_main_schedule(xlsx_bytes: bytes, group_query: str, target_date: date) -> dict | None:
    """
    Парсит основной файл расписания корпуса 2.
    Находит нужную группу и день недели, возвращает пары.
    """
    from parser import _group_matches

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error("Ошибка открытия основного файла: %s", e)
        return None

    week = _get_week_number(target_date)
    day_of_week = target_date.weekday()  # 0=пн, 5=сб
    target_day_name = DAYS_RU_CAP.get(day_of_week, '')

    # 1. Находим столбец группы
    group_col = None
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val:
            name = str(val).replace('\n', ' ').strip()
            if _group_matches(name, group_query):
                group_col = c
                break

    if group_col is None:
        logger.warning("Группа '%s' не найдена в основном файле", group_query)
        return None

    teacher_col = group_col + 1
    room_col    = group_col + 2

    # 2. Находим строки нужного дня
    day_start = day_end = None
    for rng in ws.merged_cells.ranges:
        val = ws.cell(rng.min_row, rng.min_col).value
        if val and str(val).strip().upper() == target_day_name.upper():
            day_start = rng.min_row
            day_end   = rng.max_row
            break

    if day_start is None:
        # Ищем без объединения
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val and str(val).strip().upper() == target_day_name.upper():
                day_start = r
                day_end   = r + 6
                break

    if day_start is None:
        logger.warning("День '%s' не найден в файле", target_day_name)
        return None

    # 3. Читаем пары
    pairs = []
    for r in range(day_start, day_end + 1):
        raw_num = _fmt_val(ws.cell(r, 2).value)
        if not raw_num:
            continue

        subj_val    = ws.cell(r, group_col).value
        teacher_val = ws.cell(r, teacher_col).value
        room_val    = ws.cell(r, room_col).value

        # Обрабатываем НЕД-паттерн в предмете
        subj_str, _, _ = _parse_ned_cell(subj_val, week) if subj_val and 'НЕД' in str(subj_val).upper() else (_fmt_val(subj_val), '', '')

        if not subj_str or subj_str.upper() in ('НЕТ', 'НЕТ.', '-', ''):
            continue

        teacher = _fmt_val(teacher_val)
        room    = _fmt_val(room_val)

        # Разбиваем многострочные значения (несколько преподавателей)
        teacher = ' / '.join([t.strip() for t in re.split(r'\n', teacher) if t.strip()])
        room    = ' / '.join([t.strip() for t in re.split(r'\n', room) if t.strip()])

        # Разбиваем "1,2 МДК" на отдельные пары
        for num in _split_multi_pair(raw_num):
            if num == '0':
                continue
            pairs.append({
                'num':     num,
                'subject': subj_str,
                'teacher': teacher,
                'room':    room,
            })

    return {
        'date':      target_date.strftime('%d.%m.%Y'),
        'day':       target_day_name,
        'group':     group_query,
        'corp_name': '2 корпус',
        'pairs':     pairs,
        'week':      week,
        'source':    'main',
    }


# ─── Парсер файла замен ───────────────────────────────────────────────────────

def parse_replacements(xlsx_bytes: bytes, group_query: str) -> dict | None:
    """
    Парсит файл замен корпуса 2 (вертикальная таблица).
    Возвращает пары которые нужно заменить/добавить.
    """
    from parser import _group_matches, _fmt, _is_group_header, _extract_date, _extract_day

    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error("Ошибка открытия файла замен: %s", e)
        return None

    max_row = ws.max_row

    # Дата и день из строки 1-3
    file_date = None
    day_str   = ''
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val:
                s = str(val)
                if not file_date:
                    file_date = _extract_date(s)
                if not day_str:
                    day_str = _extract_day(s)

    # Ищем группу
    start_row = end_row = None
    for r in range(1, max_row + 1):
        name = _is_group_header(ws, r, min_span=2)
        if name is None:
            continue
        if _group_matches(name, group_query):
            start_row = r
            for nr in range(r + 1, max_row + 2):
                if nr > max_row:
                    end_row = max_row
                    break
                if _is_group_header(ws, nr, min_span=2) is not None:
                    end_row = nr - 1
                    break
            break

    if start_row is None:
        return None  # группы нет в заменах — это нормально

    # Читаем пары замен
    pairs = []
    for r in range(start_row + 1, end_row + 1):
        raw_num = ws.cell(r, 1).value
        if not raw_num:
            continue
        m = re.match(r'^(\d+)', str(raw_num).strip())
        if not m:
            continue
        num_str = m.group(1)

        subject = _fmt(ws.cell(r, 2).value)
        teacher = _fmt(ws.cell(r, 3).value)
        room    = _fmt(ws.cell(r, 4).value)

        if not subject and not teacher:
            continue

        # Разбиваем многопарные записи
        for num in _split_multi_pair(num_str):
            pairs.append({
                'num':     num,
                'subject': subject or '—',
                'teacher': teacher,
                'room':    room,
            })

    return {
        'date':  file_date.strftime('%d.%m.%Y') if file_date else '',
        'day':   day_str,
        'pairs': pairs,
    }


# ─── Объединение основного расписания и замен ─────────────────────────────────

def merge_with_replacements(main: dict, replacements: dict | None) -> dict:
    """
    Накладывает замены на основное расписание.
    Замены дополняют/заменяют конкретные пары по номеру.
    """
    if not replacements or not replacements.get('pairs'):
        return main

    # Индексируем основные пары по номеру
    main_pairs = {p['num']: p for p in main.get('pairs', [])}

    for rep_pair in replacements['pairs']:
        num = rep_pair['num']
        if rep_pair['subject'] and rep_pair['subject'].upper() not in ('НЕТ', '-', ''):
            main_pairs[num] = rep_pair
        elif num in main_pairs:
            # "НЕТ" в заменах означает отмену пары
            del main_pairs[num]

    # Восстанавливаем отсортированный список
    try:
        sorted_pairs = sorted(main_pairs.values(), key=lambda p: int(p['num']))
    except (ValueError, TypeError):
        sorted_pairs = list(main_pairs.values())

    result = dict(main)
    result['pairs'] = sorted_pairs
    if replacements.get('date'):
        result['date'] = replacements['date']
    return result
