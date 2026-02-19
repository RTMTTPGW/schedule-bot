import os
import io
import requests
from datetime import datetime, timedelta
from uuid import uuid4

from aiogram import Bot, Dispatcher, types
from aiogram.types import InlineQueryResultArticle, InputTextMessageContent
from aiogram.webhook.aiohttp_server import SimpleRequestHandler, setup_application
from aiohttp import web

from openpyxl import load_workbook

TOKEN = os.getenv("BOT_TOKEN")
DRIVE_API_KEY = os.getenv("DRIVE_API_KEY")

FOLDER_ID = "1fxehYVWNrEC5EoHnrzgaxoSyCDXCDTur"
GROUP_NAME = "2-24 ОРП-1"

if not TOKEN:
    raise ValueError("BOT_TOKEN не найден")

if not DRIVE_API_KEY:
    raise ValueError("DRIVE_API_KEY не найден")


# ================= GOOGLE DRIVE API =================

def find_file_id():
    today = datetime.now()
    dates = [
        today + timedelta(days=1),
        today,
        today - timedelta(days=1),
    ]

    url = "https://www.googleapis.com/drive/v3/files"

    params = {
        "q": f"'{FOLDER_ID}' in parents",
        "fields": "files(id, name)",
        "key": DRIVE_API_KEY
    }

    response = requests.get(url, params=params)
    data = response.json()

    files = data.get("files", [])

    for date in dates:
        date_str = date.strftime("%d.%m.%Y")

        for file in files:
            if date_str in file["name"]:
                return file["id"]

    return None


def download_file(file_id):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}/export"

    params = {
        "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "key": DRIVE_API_KEY
    }

    response = requests.get(url, params=params)

    if response.status_code != 200:
        raise Exception(f"Ошибка загрузки файла: {response.text}")

    return io.BytesIO(response.content)


# ================= PARSE XLSX =================

def parse_schedule(file_bytes):
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(file_bytes, data_only=True)
    sheet = wb.active
    
    # Заполняем объединённые ячейки (очень важно!)
    for merged in list(sheet.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged.bounds
        value = sheet.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                sheet.cell(r, c).value = value

    schedule = []
    group_found = False
    target_group = GROUP_NAME.strip().lower()  # "2-24 орп-1"

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row is None:
            continue
            
        row_values = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # Ищем строку с названием группы (может быть в любой ячейке)
        if not group_found:
            row_text = " ".join(row_values).lower()
            if target_group in row_text:
                group_found = True
                # Можно сразу продолжить — группа часто в той же строке, что и заголовок
                continue
        
        if not group_found:
            continue
        
        # Номер пары — всегда в столбце A (индекс 0)
        first_cell = row_values[0]
        if not first_cell:
            continue  # пустая строка — пропускаем
        
        # Пытаемся понять, что это номер пары
        num_str = str(first_cell).strip().rstrip('.,)').lstrip(' (')
        if not num_str.isdigit():
            # если не число — возможно заголовок дня или пусто → пропускаем, но не break
            continue
        
        pair_num = num_str
        
        # Сдвиг столбцов под твою таблицу
        subject   = row_values[1] if len(row_values) > 1 else ""   # B — предмет
        teacher   = row_values[3] if len(row_values) > 3 else ""   # D — преподаватель
        cabinet   = row_values[5] if len(row_values) > 5 else ""   # F — кабинет
        
        # Чистим от лишнего
        subject = subject.replace('\n', ' ').strip()
        teacher = teacher.replace('\n', ' ').strip()
        cabinet = cabinet.replace('\n', ' ').strip()
        
        if subject or teacher or cabinet:  # если хоть что-то есть
            line = f"{pair_num}. {subject}"
            if teacher:
                line += f"\nПреп: {teacher}"
            if cabinet:
                line += f"\nКаб: {cabinet}"
            schedule.append(line)

    if not schedule:
        return f"Расписание для {GROUP_NAME} не найдено или таблица имеет другую структуру."
    
    return "\n\n".join(schedule)  # двойной перенос для красоты

# ================= TELEGRAM =================

bot = Bot(token=TOKEN)
dp = Dispatcher()


@dp.inline_query()
async def inline_handler(inline_query: types.InlineQuery):
    try:
        file_id = find_file_id()

        if not file_id:
            text = "Файл расписания не найден."
        else:
            file_bytes = download_file(file_id)
            text = parse_schedule(file_bytes)

    except Exception as e:
        text = f"Ошибка: {str(e)}"

    result = InlineQueryResultArticle(
        id=str(uuid4()),
        title="Расписание 2-24 ОРП-1",
        input_message_content=InputTextMessageContent(message_text=text)
    )

    await inline_query.answer([result], cache_time=1)


# ================= WEBHOOK =================

async def on_startup(app):
    webhook_url = os.getenv("RENDER_EXTERNAL_URL")
    await bot.set_webhook(webhook_url)


app = web.Application()
app.on_startup.append(on_startup)

SimpleRequestHandler(dispatcher=dp, bot=bot).register(app, path="/")
setup_application(app, dp, bot=bot)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    web.run_app(app, host="0.0.0.0", port=port)
