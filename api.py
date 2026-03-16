"""
api.py — REST API для сайта расписания ВПТ.

Эндпоинты:
  GET /corps                        — список корпусов
  GET /groups?corp=corp3            — список групп из последнего файла корпуса
  GET /schedule?corp=corp3&group=2-24 ОРП-1  — расписание для группы
"""

import io
import re
import logging
from datetime import date

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
import openpyxl

from config import CORPS, CORPS_BY_ID, get_current_semester
from drive import get_files_for_corp, export_as_xlsx
from parser import parse_file
from sheets import get_today_file_id, get_latest_file_id, parse_schedule

logger = logging.getLogger(__name__)

app = FastAPI(
    title="ВПТ Расписание API",
    description="API для получения расписания Волжского политехнического техникума",
    version="1.0.0",
)

# Разрешаем запросы с любого домена (нужно для GitHub Pages)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET"],
    allow_headers=["*"],
)


# ─── /corps ───────────────────────────────────────────────────────────────────

@app.get("/corps")
def get_corps():
    """Возвращает список корпусов."""
    return [
        {"id": c["id"], "name": c["name"]}
        for c in CORPS
    ]


# ─── /groups ──────────────────────────────────────────────────────────────────

def _extract_groups_from_file(xlsx_bytes: bytes, table_format: str) -> list[str]:
    """Извлекает список групп из xlsx файла."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        groups = []
        skip = (
            "расписание", "учебная группа", "номер пары",
            "дисциплина", "преподаватель", "аудитор",
            "лист замен", "изменения", "корпус", "гб поу",
            "пара", "кабинет", "территория",
        )
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row, column=1)
            for rng in ws.merged_cells.ranges:
                if rng.min_row == row and rng.min_col == 1 and rng.max_col >= (2 if table_format == "type_c" else 4):
                    val = cell_a.value
                    if not val:
                        continue
                    val_str = str(val).strip()
                    if any(s in val_str.lower() for s in skip):
                        continue
                    if val_str and val_str not in groups:
                        groups.append(val_str)
        return groups
    except Exception as e:
        logger.error("Ошибка извлечения групп: %s", e)
        return []


@app.get("/groups")
def get_groups(corp: str = Query(..., description="ID корпуса, например corp3")):
    """Возвращает список групп из последнего файла корпуса."""
    corp_cfg = CORPS_BY_ID.get(corp)
    if not corp_cfg:
        raise HTTPException(status_code=404, detail=f"Корпус '{corp}' не найден")

    try:
        # Для корпуса 2 берём группы из основного файла расписания
        if corp == "corp2":
            from sheets import _get_corp2_main_file
            xlsx = _get_corp2_main_file()
            if not xlsx:
                return {"corp": corp, "groups": []}
            groups = _extract_groups_from_corp2_main(xlsx)
            return {"corp": corp, "groups": groups}

        file_id = get_latest_file_id(corp)
        if not file_id:
            return {"corp": corp, "groups": []}
        xlsx = export_as_xlsx(file_id)
        groups = _extract_groups_from_file(xlsx, corp_cfg["table_format"])
        return {"corp": corp, "groups": groups}
    except Exception as e:
        logger.exception("Ошибка /groups")
        raise HTTPException(status_code=500, detail=str(e))


def _extract_groups_from_corp2_main(xlsx_bytes: bytes) -> list[str]:
    """Извлекает список групп из горизонтальной таблицы корпуса 2."""
    import io, re
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        groups = []
        skip = ('день недели', 'пара')
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if not val:
                continue
            name = str(val).replace('\n', ' ').strip()
            if any(s in name.lower() for s in skip):
                continue
            # Извлекаем короткое название — только код группы
            m = re.match(r'^(\S+-\d+\s+\S+-\d+)', name)
            short = m.group(1).strip() if m else name.split('\n')[0].strip()
            if short and short not in groups:
                groups.append(short)
        return groups
    except Exception as e:
        logger.error("Ошибка извлечения групп корпуса 2: %s", e)
        return []


# ─── /schedule ────────────────────────────────────────────────────────────────

@app.get("/schedule")
def get_schedule(
    corp:  str = Query(..., description="ID корпуса, например corp3"),
    group: str = Query(..., description="Название группы, например 2-24 ОРП-1"),
    mode:  str = Query("today", description="today — сегодня, latest — последний файл"),
):
    """
    Возвращает расписание для группы.
    mode=today  — последний файл с датой <= сегодня
    mode=latest — самый последний файл (может быть на будущий день)
    """
    corp_cfg = CORPS_BY_ID.get(corp)
    if not corp_cfg:
        raise HTTPException(status_code=404, detail=f"Корпус '{corp}' не найден")

    try:
        file_id = get_today_file_id(corp) if mode == "today" else get_latest_file_id(corp)
        if not file_id:
            raise HTTPException(status_code=404, detail="Файлов расписания не найдено")

        data = parse_schedule(file_id, group, corp)
        if not data:
            raise HTTPException(
                status_code=404,
                detail=f"Группа '{group}' не найдена в файле"
            )

        return {
            "corp":   corp,
            "corp_name": corp_cfg["name"],
            "group":  group,
            "date":   data.get("date", ""),
            "day":    data.get("day", ""),
            "pairs":  data.get("pairs", []),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка /schedule")
        raise HTTPException(status_code=500, detail=str(e))


# ─── healthcheck ──────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok"}
