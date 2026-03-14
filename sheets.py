"""
sheets.py — скачивание xlsx из публичной папки Google Drive и парсинг расписания.

Структура файла (выяснена по скриншоту):
  Строка 1:  "Расписание на 16.03.2026 Понедельник верх"  (дата и день)
  Строка 3:  "Учебная группа"  (заголовок — пропускаем)
  Строка 4:  "Номер пары | Дисциплина | Преподаватель | Аудитория"  (пропускаем)
  Строка 5:  "1-25 РЗ-8"  — название группы (объединённая ячейка A:G)
  Строки 6+: пары группы: A=номер, B:D=дисциплина, E:F=преподаватель, G=аудитория
  Строка N:  следующая группа — снова объединённая ячейка
  ...и так далее вниз

Никакого Service Account не нужно — папка публичная, читаем напрямую.
"""

import os
import io
import re
import logging
import requests
import openpyxl

logger = logging.getLogger(__name__)

FOLDER_ID = os.environ["FOLDER_ID"]
REQUEST_TIMEOUT = 30


# ─── Google Drive: список файлов и скачивание ─────────────────────────────────

def get_drive_files() -> list[dict]:
    """
    Получает список файлов из публичной папки Google Drive.
    Возвращает [{"id": "...", "name": "..."}, ...].
    """
    url = f"https://drive.google.com/drive/folders/{FOLDER_ID}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    resp = requests.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()

    # Drive встраивает file id в JS — они всегда 33 символа, начинаются с 1
    file_ids = re.findall(r'"(1[a-zA-Z0-9_-]{32})"', resp.text)
    file_ids = list(dict.fromkeys(file_ids))  # убираем дубли, сохраняем порядок

    logger.info("Найдено файлов в папке: %d", len(file_ids))
    return [{"id": fid, "name": ""} for fid in file_ids]


def get_latest_file_id() -> str | None:
    """Возвращает id последнего добавленного файла в папке."""
    files = get_drive_files()
    return files[0]["id"] if files else None


def download_xlsx(file_id: str) -> bytes:
    """Скачивает xlsx-файл по его Google Drive file_id."""
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        )
    }
    session = requests.Session()
    resp = session.get(url, headers=headers, timeout=REQUEST_TIMEOUT)

    # Google может показать страницу подтверждения для больших файлов
    if b"confirm" in resp.content and b"virus scan" in resp.content.lower():
        token = re.search(r'confirm=([0-9A-Za-z_]+)', resp.text)
        if token:
            resp = session.get(
                f"{url}&confirm={token.group(1)}",
                headers=headers,
                timeout=REQUEST_TIMEOUT,
            )

    resp.raise_for_status()
    logger.info("Скачан файл id=%s, %d байт", file_id, len(resp.content))
    return resp.content


# ─── Парсинг xlsx ─────────────────────────────────────────────────────────────

def _cell_val(ws, row: int, col: int) -> str:
    """
    Читает значение ячейки с учётом объединённых ячеек.
    Если ячейка — часть merged range, возвращает значение левой верхней ячейки.
    """
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            val = ws.cell(rng.min_row, rng.min_col).value
            return str(val).strip() if val is not None else ""
    val = cell.value
    return str(val).strip() if val is not None else ""


def _detect_group_row(ws, row: int) -> str | None:
    """
    Определяет, является ли строка заголовком группы.
    Признак: ячейка A объединена с соседними (минимум 5 столбцов),
    и значение не является системным заголовком таблицы.
    Возвращает название группы или None.
    """
    cell_a = ws.cell(row=row, column=1)
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row and rng.min_col == 1 and rng.max_col >= 5:
            val = cell_a.value
            if not val:
                continue
            val_str = str(val).strip()
            skip = ("расписание", "учебная группа", "номер пары",
                    "дисциплина", "преподаватель", "аудитор")
            if any(s in val_str.lower() for s in skip):
                return None
            if val_str:
                return val_str
    return None


def parse_schedule(file_id: str, group_name: str) -> dict | None:
    """
    Скачивает xlsx и возвращает расписание для группы:
    {
        "date":  "16.03.2026",
        "day":   "Понедельник",
        "group": "2-24 ОРП-1",
        "pairs": [
            {"num": "1", "subject": "...", "teacher": "...", "room": "..."},
            ...
        ]
    }
    Возвращает None если группа не найдена.
    """
    data = download_xlsx(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    max_row = ws.max_row

    # ── Дата и день из строки 1 ──────────────────────────────────────────────
    row1 = _cell_val(ws, 1, 1)
    # "Расписание на 16.03.2026 Понедельник верх"
    date_str = ""
    day_str = ""
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})', row1)
    if m:
        date_str = m.group(1)
    m = re.search(
        r'(понедельник|вторник|среда|четверг|пятница|суббота|воскресенье)',
        row1.lower()
    )
    if m:
        day_str = m.group(1).capitalize()

    # ── Находим блок нужной группы ───────────────────────────────────────────
    group_clean = group_name.strip().lower()
    start_row = end_row = None

    for r in range(1, max_row + 1):
        name = _detect_group_row(ws, r)
        if name is None:
            continue
        if group_clean in name.lower():
            start_row = r
            # Конец блока — следующий заголовок группы или конец листа
            for nr in range(r + 1, max_row + 2):
                if nr > max_row:
                    end_row = max_row
                    break
                if _detect_group_row(ws, nr) is not None:
                    end_row = nr - 1
                    break
            break

    if start_row is None:
        logger.warning("Группа «%s» не найдена в файле %s", group_name, file_id)
        return None

    logger.info("Группа «%s» — строки %d:%d", group_name, start_row, end_row)

    # ── Читаем пары ──────────────────────────────────────────────────────────
    pairs = []
    for r in range(start_row + 1, end_row + 1):
        num = _cell_val(ws, r, 1)      # A  — номер пары
        if not num:
            continue

        subject = _cell_val(ws, r, 2)  # B  (объединение B:D)
        teacher = _cell_val(ws, r, 5)  # E  (объединение E:F)
        room    = _cell_val(ws, r, 7)  # G  — аудитория

        if not subject and not teacher:
            continue

        pairs.append({
            "num":     num,
            "subject": subject or "—",
            "teacher": teacher,
            "room":    room,
        })

    return {
        "date":  date_str,
        "day":   day_str,
        "group": group_name,
        "pairs": pairs,
    }


# ─── Форматирование ───────────────────────────────────────────────────────────

# Время пар — поменяй если у вас другое
PAIR_TIMES = {
    "1": "08:00–09:35",
    "2": "09:45–11:20",
    "3": "11:40–13:15",
    "4": "13:25–15:00",
    "5": "15:20–16:55",
    "6": "17:05–18:40",
    "7": "18:50–20:25",
}


def format_schedule(data: dict) -> str:
    """Форматирует расписание в HTML-текст для Telegram."""
    date  = data.get("date", "")
    day   = data.get("day", "")
    group = data.get("group", "")
    pairs = data.get("pairs", [])

    header = f"📅 <b>{date}"
    if day:
        header += f", {day}"
    header += f"</b>\n👥 Группа: <b>{group}</b>\n"
    header += "━━━━━━━━━━━━━━━━━━"

    if not pairs:
        return header + "\n\n🎉 Занятий нет!"

    lines = [header]
    for p in pairs:
        time = PAIR_TIMES.get(str(p["num"]), "")
        time_str = f"  🕐 <i>{time}</i>" if time else ""
        block = f"\n🔹 <b>{p['num']} пара</b>{time_str}\n   📖 {p['subject']}"
        if p["teacher"]:
            block += f"\n   👩‍🏫 {p['teacher']}"
        if p["room"]:
            block += f"\n   🏫 {p['room']}"
        lines.append(block)

    return "\n".join(lines)
