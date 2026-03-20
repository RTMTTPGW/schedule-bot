"""
parser.py — универсальный парсер таблиц расписания.

Поддерживает три формата:
  type_a — корпус 3: "Расписание на DD.MM.YYYY День"
           A=номер пары, B:D=дисциплина, E:F=преподаватель, G=аудитория
           Группа — объединённая ячейка на всю ширину
  type_b — корпуса 1, 4: "Лист замен на DD.MM.YYYY День"
           Та же структура что type_a, может быть столбец Территория (H)
  type_c — корпус 2: строки 1-3 заголовок с датой
           A=пара (текст "1пара"/"1 пара"), B=дисциплина, C=преподаватель, D=кабинет
           Группа — объединённая ячейка, название может быть слитным
"""

import io
import re
import logging
import openpyxl
from datetime import date

logger = logging.getLogger(__name__)

DAYS_RU = {
    "понедельник", "вторник", "среда", "четверг",
    "пятница", "суббота", "воскресенье"
}

# ─── Хелперы ──────────────────────────────────────────────────────────────────

def _fmt(val) -> str:
    """Форматирует значение ячейки. float без дробной части → int."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _split_multi_pair_num(num_str: str) -> list[str]:
    """
    Разбивает номер пары на список отдельных номеров.

    Поддерживаемые форматы:
      "1"       → ['1']
      "1,2"     → ['1', '2']
      "1-2"     → ['1', '2']
      "1-3"     → ['1', '2', '3']   ← диапазон
      "1,2,3"   → ['1', '2', '3']
      "1, 2, 3" → ['1', '2', '3']
      "2,4"     → ['2', '4']        ← перечисление без диапазона
    """
    s = str(num_str).strip()

    # Диапазон: "1-3", "2-4" и т.д.
    m_range = re.match(r'^(\d+)\s*[-–]\s*(\d+)$', s)
    if m_range:
        start, end = int(m_range.group(1)), int(m_range.group(2))
        if start < end and (end - start) <= 5:  # защита от мусора вроде "1-99"
            return [str(i) for i in range(start, end + 1)]

    # Перечисление через запятую: "1,2,3" или "1, 2, 3"
    if re.match(r'^\d+(\s*,\s*\d+)+$', s):
        return [p.strip() for p in s.split(',')]

    # Одиночный номер
    return [s]


def _cell(ws, row: int, col: int) -> str:
    """Читает ячейку с учётом объединений."""
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            val = ws.cell(rng.min_row, rng.min_col).value
            return _fmt(val)
    return _fmt(cell.value)


def _extract_date(text: str) -> date | None:
    """Извлекает дату из строки вида DD.MM.YYYY."""
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', text)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def _extract_day(text: str) -> str:
    """Извлекает день недели из строки."""
    for day in DAYS_RU:
        if day in text.lower():
            return day.capitalize()
    return ""


def _is_group_header(ws, row: int, min_span: int = 4) -> str | None:
    """
    Проверяет является ли строка заголовком группы.
    Признак: ячейка A объединена с соседними (min_span столбцов),
    значение не системный заголовок.
    """
    cell_a = ws.cell(row=row, column=1)
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row and rng.min_col == 1 and rng.max_col >= min_span:
            val = cell_a.value
            if not val:
                continue
            val_str = str(val).strip()
            skip = (
                "расписание", "учебная группа", "номер пары",
                "дисциплина", "преподаватель", "аудитор",
                "лист замен", "изменения", "корпус", "гб поу",
                "пара", "кабинет", "территория",
            )
            if any(s in val_str.lower() for s in skip):
                return None
            if val_str:
                return val_str
    return None


def _is_valid_group_query(query: str) -> bool:
    """
    Проверяет что запрос достаточно конкретный чтобы искать группу.
    Минимальные требования:
    - длина >= 4 символов
    - содержит хотя бы одну цифру
    - содержит хотя бы одну букву
    - не состоит только из цифр или только из букв
    """
    q = query.strip()
    q = q.strip()
    if len(q) < 4:
        return False
    has_digit = any(c.isdigit() for c in q)
    has_alpha = any(c.isalpha() for c in q)
    if not (has_digit and has_alpha):
        return False
    # Должен содержать дефис или начинаться с цифры (формат групп: "2-24 ОРП-1", "ОРП-1")
    starts_with_digit = q[0].isdigit()
    has_hyphen = '-' in q
    return starts_with_digit or has_hyphen


def _group_matches(group_header: str, group_query: str) -> bool:
    """
    Проверяет совпадение названия группы.
    Стратегии (применяются по порядку, первое совпадение = match):
    1. Точное совпадение после нормализации пробелов
    2. Запрос является началом заголовка (корпус 2: "1-25 ПКД-10" в "1-25 ПКД-10Поварское...")
    3. Запрос без пробелов входит в начало заголовка без пробелов (слитное написание)
    """
    # Сначала проверяем что запрос вообще похож на название группы
    if not _is_valid_group_query(group_query):
        return False

    def norm_spaces(s: str) -> str:
        return re.sub(r'\s+', ' ', s).strip().lower()

    def norm_nospaces(s: str) -> str:
        return re.sub(r'\s+', '', s).lower()

    h = norm_spaces(group_header)
    q = norm_spaces(group_query)

    # 1. Точное совпадение
    if h == q:
        return True

    # 2. Заголовок начинается с запроса
    if h.startswith(q):
        rest = h[len(q):]
        if not rest or not rest[0].isalnum():
            return True

    # 3. Без пробелов — запрос входит в начало заголовка
    hn = norm_nospaces(group_header)
    qn = norm_nospaces(group_query)
    if hn.startswith(qn):
        rest = hn[len(qn):]
        if not rest or not rest[0].isdigit():
            return True

    return False


# ─── Парсер type_a / type_b ───────────────────────────────────────────────────

def _parse_type_ab(ws, group_query: str) -> dict | None:
    """
    Парсит таблицы корпусов 1, 3, 4.
    Строка 1: "Расписание на DD.MM.YYYY День" или "Лист замен на DD.MM.YYYY День"
    Группы — объединённые ячейки на всю ширину.
    A=номер, B:D=предмет, E:F=преподаватель, G=аудитория
    """
    max_row = ws.max_row

    # Дата и день из строки 1
    row1 = _cell(ws, 1, 1)
    file_date = _extract_date(row1)
    day_str = _extract_day(row1)

    # Ищем группу
    start_row = end_row = None
    for r in range(1, max_row + 1):
        name = _is_group_header(ws, r, min_span=4)
        if name is None:
            continue
        if _group_matches(name, group_query):
            start_row = r
            for nr in range(r + 1, max_row + 2):
                if nr > max_row:
                    end_row = max_row
                    break
                if _is_group_header(ws, nr, min_span=4) is not None:
                    end_row = nr - 1
                    break
            break

    if start_row is None:
        return None

    # Читаем пары
    pairs = []
    for r in range(start_row + 1, end_row + 1):
        raw_num = _cell(ws, r, 1)
        if not raw_num:
            continue

        # Нормализуем номер пары:
        # - обычное число: "1", "2.0" → "1", "2"
        # - диапазон/перечисление: "1-3", "1,2" → разворачиваем ниже
        num_str = raw_num.strip()
        try:
            # Обычное число (в т.ч. "1.0" из Excel float)
            num_str = str(int(float(num_str)))
        except ValueError:
            # Не число — может быть диапазон "1-3" или "1,2"
            # Проверяем что хотя бы начинается с цифры, иначе пропускаем
            if not re.match(r'^\d', num_str):
                continue

        subject = _cell(ws, r, 2)
        teacher = _cell(ws, r, 5)
        room = _cell(ws, r, 7)

        if not subject and not teacher:
            continue

        # Разворачиваем диапазон/перечисление в отдельные пары
        for split_num in _split_multi_pair_num(num_str):
            pairs.append({
                "num": split_num,
                "subject": subject or "—",
                "teacher": teacher,
                "room": room,
            })

    return {
        "date": file_date.strftime("%d.%m.%Y") if file_date else "",
        "day": day_str,
        "group": group_query,
        "pairs": pairs,
    }


# ─── Парсер type_c ────────────────────────────────────────────────────────────

def _parse_type_c(ws, group_query: str) -> dict | None:
    """
    Парсит таблицы корпуса 2.
    Строки 1-3: заголовок с датой ("Изменения в расписании на DD.MM.YYYY")
    A=пара ("1пара"/"1 пара"), B=предмет, C=преподаватель, D=кабинет
    Группа — объединённая ячейка (название может быть слитным: "1-25ПКД-10Поварское...")
    """
    max_row = ws.max_row

    # Дата из первых трёх строк
    file_date = None
    day_str = ""
    for r in range(1, 4):
        val = _cell(ws, r, 1)
        if not val:
            # Попробуем другие столбцы
            for c in range(1, 5):
                val = _cell(ws, r, c)
                if val:
                    break
        if val:
            if not file_date:
                file_date = _extract_date(val)
            if not day_str:
                day_str = _extract_day(val)

    # Ищем группу — объединённая ячейка, может быть слитной
    start_row = end_row = None
    for r in range(1, max_row + 1):
        name = _is_group_header(ws, r, min_span=2)
        if name is None:
            continue
        if _group_matches(name, group_query):
            start_row = r
            for nr in range(r + 1, max_row + 2):
                if nr > max_row:
                    end_row = max_row
                    break
                if _is_group_header(ws, nr, min_span=2) is not None:
                    end_row = nr - 1
                    break
            break

    if start_row is None:
        return None

    # Читаем пары
    pairs = []
    for r in range(start_row + 1, end_row + 1):
        raw_num = _cell(ws, r, 1)
        if not raw_num:
            continue

        # Нормализуем "1пара" / "1 пара" / "1-2 пара" / "1" → num_str
        # Сначала вытаскиваем всё числовое выражение до слова "пара"
        m = re.match(r'^([\d,\s\-–]+?)(?:\s*пара)?$', raw_num.strip(), re.IGNORECASE)
        if not m:
            continue
        num_str = m.group(1).strip()
        if not num_str:
            continue

        subject = _cell(ws, r, 2)
        teacher = _cell(ws, r, 3)
        room = _cell(ws, r, 4)

        if not subject and not teacher:
            continue

        for split_num in _split_multi_pair_num(num_str):
            pairs.append({
                "num": split_num,
                "subject": subject or "—",
                "teacher": teacher,
                "room": room,
            })

    return {
        "date": file_date.strftime("%d.%m.%Y") if file_date else "",
        "day": day_str,
        "group": group_query,
        "pairs": pairs,
    }


# ─── Парсер type_d (корпус 2: горизонтальная таблица + замены) ───────────────

DAYS_RU_LIST = ['ПОНЕДЕЛЬНИК', 'ВТОРНИК', 'СРЕДА', 'ЧЕТВЕРГ', 'ПЯТНИЦА', 'СУББОТА']


def _get_week_number(d=None) -> int:
    """1 или 2 — номер учебной недели по ISO."""
    from datetime import date as _date
    if d is None:
        d = _date.today()
    return 1 if d.isocalendar()[1] % 2 == 1 else 2


def _cv(ws, r: int, c: int) -> str:
    """Значение ячейки с учётом объединений."""
    cell = ws.cell(r, c)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return str(ws.cell(rng.min_row, rng.min_col).value or '').strip()
    return str(cell.value or '').strip()


def _first_line(s: str) -> str:
    return s.split('\n')[0].strip()


def _group_match_horiz(header: str, query: str) -> bool:
    def n(s): return re.sub(r'\s+', '', _first_line(s)).lower()
    h, q = n(header), n(query)
    if h == q: return True
    if h.startswith(q):
        rest = h[len(q):]
        return not rest or not rest[0].isalnum()
    return q in h


def _week_val(val: str, week: int) -> str:
    """Извлекает значение для нужной недели из ячейки с маркерами НЕД1/НЕД2."""
    if not val: return ''
    if val.upper().strip() in ('НЕТ',): return 'НЕТ'
    if not re.search(r'\dНЕД|\d\s*НЕД', val, re.IGNORECASE): return val
    parts = re.split(r'\n|(?=/?\s*\d\s*НЕД)', val, flags=re.IGNORECASE)
    for part in parts:
        m = re.match(r'\s*/?\s*(\d)\s*НЕД[-.\s]*(.*)', part.strip(), re.IGNORECASE)
        if m and int(m.group(1)) == week:
            r2 = m.group(2).strip().strip('-').strip('/').strip()
            return r2 or ''
    return ''


def _split_multi_pair(pn: str, subj: str, teacher: str, room: str) -> list:
    """Разбивает '1,2 МДК' на две отдельные пары."""
    cm = re.match(r'^(\d+)[,\s]+(\d+)\s+(.+)$', subj)
    if cm:
        return [
            {'num': cm.group(1), 'subject': cm.group(3).strip(), 'teacher': teacher, 'room': room},
            {'num': cm.group(2), 'subject': cm.group(3).strip(), 'teacher': teacher, 'room': room},
        ]
    return [{'num': pn, 'subject': subj, 'teacher': teacher, 'room': room}]


def _parse_corp2_horizontal(ws, group_query: str, target_date) -> dict | None:
    """Парсит основную горизонтальную таблицу корпуса 2."""
    from datetime import date as _date
    week = _get_week_number(target_date)
    day_idx = target_date.weekday()
    if day_idx > 5: return None
    day_name = DAYS_RU_LIST[day_idx]

    group_col = None
    for c in range(1, ws.max_column + 1):
        val = _cv(ws, 1, c)
        if val and _group_match_horiz(val, group_query):
            group_col = c
            break
    if group_col is None: return None

    day_row = None
    for r in range(1, ws.max_row + 1):
        if _cv(ws, r, 1).upper() == day_name:
            day_row = r
            break
    if day_row is None: return None

    pairs = []
    for offset in range(0, 8):
        r = day_row + offset
        pn_raw = _cv(ws, r, 2)
        try:
            pn = str(int(float(pn_raw)))
        except (ValueError, TypeError):
            continue

        subj_raw = _cv(ws, r, group_col)
        teacher = _first_line(_cv(ws, r, group_col + 1))
        room = _first_line(_cv(ws, r, group_col + 2))
        subj = _week_val(subj_raw, week)
        if not subj or subj.upper() in ('НЕТ', '-', ''):
            continue

        pairs.extend(_split_multi_pair(pn, subj, teacher, room))

    return {
        'date': target_date.strftime('%d.%m.%Y'),
        'day': day_name.capitalize(),
        'group': group_query,
        'pairs': pairs,
        'week': week,
    }


def _parse_corp2_substitutions(ws, group_query: str) -> dict:
    """Парсит файл замен, возвращает {pair_num: pair_data}."""
    def n(s): return re.sub(r'\s+', '', str(s)).lower()
    q = n(group_query)

    in_group = False
    subs = {}
    for r in range(1, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        c2 = str(ws.cell(r, 2).value or '').strip()
        c3 = str(ws.cell(r, 3).value or '').strip()
        c4 = str(ws.cell(r, 4).value or '').strip()

        # Заголовок группы
        if (not c1 or c1 in (' ',)) and c2 and c2 not in ('Дисциплина', ' '):
            h = n(c2)
            if q in h or h.startswith(q):
                in_group = True
                subs = {}
            elif in_group:
                break
            else:
                in_group = False
            continue

        if not in_group: continue

        m = re.match(r'(\d+)', c1)
        if not m: continue
        pn = m.group(1)
        if c2 and c2 not in (' ',):
            subs[pn] = {'subject': c2, 'teacher': c3, 'room': c4}

    return subs


def _parse_type_d(main_ws, subs_ws_list: list, group_query: str, target_date) -> dict | None:
    """
    Полный парсер корпуса 2:
    1. Берёт базовое расписание из горизонтальной таблицы
    2. Накладывает замены из файлов замен (если есть на нужную дату)
    """
    result = _parse_corp2_horizontal(main_ws, group_query, target_date)
    if result is None: return None

    for subs_ws in subs_ws_list:
        subs = _parse_corp2_substitutions(subs_ws, group_query)
        if subs:
            base = {p['num']: p for p in result['pairs']}
            for pn, sub in subs.items():
                if sub['subject'].lower() in ('нет', ''):
                    base.pop(pn, None)
                else:
                    base[pn] = {'num': pn, **sub}
            result['pairs'] = sorted(base.values(), key=lambda x: int(x['num']))

    return result


# ─── Публичный интерфейс ──────────────────────────────────────────────────────

def parse_file(xlsx_bytes: bytes, table_format: str, group_query: str,
               subs_xlsx_list: list | None = None, target_date=None) -> dict | None:
    """
    Парсит xlsx-файл и возвращает расписание для группы.

    table_format: "type_a", "type_b", "type_c", "type_d"
    subs_xlsx_list: список bytes файлов замен (только для type_d)
    target_date: дата для которой ищем расписание (только для type_d)
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error("Ошибка открытия файла: %s", e)
        return None

    if table_format in ("type_a", "type_b"):
        return _parse_type_ab(ws, group_query)
    elif table_format == "type_c":
        return _parse_type_c(ws, group_query)
    elif table_format == "type_d":
        from datetime import date as _date
        td = target_date or _date.today()
        subs_ws_list = []
        for subs_bytes in (subs_xlsx_list or []):
            try:
                subs_wb = openpyxl.load_workbook(io.BytesIO(subs_bytes), data_only=True)
                subs_ws_list.append(subs_wb.active)
            except Exception as e:
                logger.warning("Ошибка открытия файла замен: %s", e)
        return _parse_type_d(ws, subs_ws_list, group_query, td)
    else:
        logger.warning("Неизвестный формат таблицы: %s", table_format)
        return None


def get_file_date(xlsx_bytes: bytes, table_format: str) -> date | None:
    """Извлекает дату из файла без полного парсинга."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), data_only=True, read_only=True
        )
        ws = wb.active
        # Для всех форматов дата в первых 3 строках
        for r in range(1, 4):
            for row in ws.iter_rows(min_row=r, max_row=r, values_only=True):
                for cell in row:
                    if cell:
                        d = _extract_date(str(cell))
                        if d:
                            wb.close()
                            return d
        wb.close()
    except Exception as e:
        logger.warning("Ошибка чтения даты: %s", e)
    return None
